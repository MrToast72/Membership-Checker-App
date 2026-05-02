from __future__ import annotations

import re
import sqlite3
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

import openpyxl


def normalize(value: str) -> str:
    value = value or ""
    return re.sub(r"[^a-z0-9]", "", value.strip().lower())


def canonical_header(value: str) -> str:
    return normalize(value)


def split_first_name(first_name: str) -> list[str]:
    if not first_name:
        return []
    parts = [p.strip() for p in re.split(r"[/&]| and ", first_name, flags=re.IGNORECASE)]
    parts = [p for p in parts if p]
    return parts or [first_name.strip()]


def parse_yes_no(value: str) -> str:
    cleaned = (value or "").strip().lower()
    if cleaned in {"yes", "y", "true", "1"}:
        return "Yes"
    if cleaned in {"no", "n", "false", "0", ""}:
        return "No"
    return "No"


@dataclass
class MemberRecord:
    record_id: int | None
    first_name: str
    last_name: str
    email: str
    membership_type: str
    membership_number: str
    includes_cart: str
    includes_range: str
    membership_amount_used: int
    sheet_name: str | None = None
    row_number: int | None = None

    @property
    def display_name(self) -> str:
        return f"{self.first_name} {self.last_name}".strip()


@dataclass
class SheetConfig:
    sheet_name: str
    header_row: int
    index_map: dict[str, int]


class ExcelMembershipLoader:
    COL_ALIASES = {
        "first_name": {"firstname"},
        "last_name": {"lastname"},
        "email": {"email"},
        "membership_type": {"membershiptype"},
        "membership_number": {"membershipnumber", "membernumber"},
        "membership_amount_used": {"membershipamountused", "amountused", "membershipused"},
        "includes_cart": {"includescart", "includecart"},
        "includes_range": {"includesrange", "includerange"},
    }

    def load(self, file_obj) -> list[MemberRecord]:
        records: list[MemberRecord] = []
        workbook = openpyxl.load_workbook(file_obj, data_only=True)
        try:
            for sheet_name in workbook.sheetnames:
                if "total" in sheet_name.lower():
                    continue
                sheet = workbook[sheet_name]
                config = self._find_sheet_config(sheet_name, sheet)
                if not config:
                    continue
                records.extend(self._load_sheet_records(sheet_name, sheet, config))
        finally:
            workbook.close()
        return records

    def _find_sheet_config(self, sheet_name: str, sheet) -> SheetConfig | None:
        header_row = 0
        raw_headers: list[str] = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=30, values_only=True), start=1):
            cells = [str(c).strip() if c is not None else "" for c in row]
            canon = {canonical_header(c) for c in cells if c}
            if "firstname" in canon and "lastname" in canon:
                header_row = row_idx
                raw_headers = cells
                break
        if not header_row:
            return None

        source = {canonical_header(v): idx for idx, v in enumerate(raw_headers, start=1) if v}
        index_map: dict[str, int] = {}
        for logical, aliases in self.COL_ALIASES.items():
            for alias in aliases:
                if alias in source:
                    index_map[logical] = source[alias]
                    break

        required = {"first_name", "last_name", "membership_type"}
        if not required.issubset(index_map.keys()):
            return None

        return SheetConfig(sheet_name=sheet_name, header_row=header_row, index_map=index_map)

    def _cell_text(self, row, idx: int | None) -> str:
        if not idx:
            return ""
        value = row[idx - 1]
        if value is None:
            return ""
        return str(value).strip()

    def _cell_int(self, row, idx: int | None) -> int:
        if not idx:
            return 0
        value = row[idx - 1]
        if value is None or str(value).strip() == "":
            return 0
        try:
            return int(float(str(value).strip()))
        except ValueError:
            return 0

    def _load_sheet_records(self, sheet_name: str, sheet, config: SheetConfig) -> list[MemberRecord]:
        records: list[MemberRecord] = []
        for row_idx, row in enumerate(
            sheet.iter_rows(min_row=config.header_row + 1, max_row=sheet.max_row, values_only=True),
            start=config.header_row + 1,
        ):
            first_name = self._cell_text(row, config.index_map.get("first_name"))
            last_name = self._cell_text(row, config.index_map.get("last_name"))
            email = self._cell_text(row, config.index_map.get("email"))
            membership_number = self._cell_text(row, config.index_map.get("membership_number"))
            if not (first_name or last_name or email or membership_number):
                continue

            record = MemberRecord(
                record_id=None,
                first_name=first_name,
                last_name=last_name,
                email=email,
                membership_type=self._cell_text(row, config.index_map.get("membership_type")),
                membership_number=membership_number,
                includes_cart=parse_yes_no(self._cell_text(row, config.index_map.get("includes_cart")) or "No"),
                includes_range=parse_yes_no(self._cell_text(row, config.index_map.get("includes_range")) or "No"),
                membership_amount_used=self._cell_int(row, config.index_map.get("membership_amount_used")),
                sheet_name=sheet_name,
                row_number=row_idx,
            )
            records.append(record)
        return records


class MemberIndex:
    def __init__(self, records: Iterable[MemberRecord]) -> None:
        self.records = list(records)
        self.by_membership_number: dict[str, list[MemberRecord]] = {}
        self.by_email: dict[str, list[MemberRecord]] = {}
        self.by_name: dict[str, list[MemberRecord]] = {}
        self._rebuild_indexes()

    def _add_to_index(self, index: dict[str, list[MemberRecord]], key: str, record: MemberRecord) -> None:
        if not key:
            return
        index.setdefault(key, []).append(record)

    def _rebuild_indexes(self) -> None:
        self.by_membership_number.clear()
        self.by_email.clear()
        self.by_name.clear()
        for record in self.records:
            self._add_to_index(self.by_membership_number, normalize(record.membership_number), record)
            self._add_to_index(self.by_email, normalize(record.email), record)
            first_names = split_first_name(record.first_name)
            last_name = record.last_name.strip()
            for first in first_names:
                self._add_to_index(self.by_name, normalize(f"{first} {last_name}"), record)
                self._add_to_index(self.by_name, normalize(f"{last_name} {first}"), record)
            self._add_to_index(self.by_name, normalize(record.display_name), record)

    def lookup(self, scan_text: str) -> list[MemberRecord]:
        raw = (scan_text or "").strip()
        if not raw:
            return []

        key = normalize(raw)
        matches: list[MemberRecord] = []
        if key in self.by_membership_number:
            matches.extend(self.by_membership_number[key])
        if not matches and key in self.by_email:
            matches.extend(self.by_email[key])
        if not matches:
            for candidate in self._name_candidates(raw):
                ckey = normalize(candidate)
                if ckey in self.by_name:
                    matches.extend(self.by_name[ckey])
        if not matches:
            matches = self._fallback_contains(raw)

        unique: list[MemberRecord] = []
        seen = set()
        for item in matches:
            signature = item.record_id or (item.sheet_name, item.row_number, item.display_name)
            if signature in seen:
                continue
            seen.add(signature)
            unique.append(item)
        return unique

    def _name_candidates(self, raw: str) -> list[str]:
        values = [raw]
        if "," in raw:
            parts = [p.strip() for p in raw.split(",") if p.strip()]
            if len(parts) >= 2:
                values.extend([f"{parts[1]} {parts[0]}", f"{parts[0]} {parts[1]}"])
        else:
            parts = [p for p in raw.split() if p]
            if len(parts) >= 2:
                values.extend([f"{parts[0]} {parts[-1]}", f"{parts[-1]} {parts[0]}"])
        return values

    def _fallback_contains(self, raw: str) -> list[MemberRecord]:
        needle = normalize(raw)
        if not needle:
            return []
        results = []
        for record in self.records:
            haystack = normalize(
                f"{record.first_name} {record.last_name} {record.email} {record.membership_number}"
            )
            if needle in haystack:
                results.append(record)
        return results


class MembershipStore:
    def __init__(self, db_path: Path) -> None:
        self.db_path = db_path
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        self._loader = ExcelMembershipLoader()
        self._index = MemberIndex([])
        self._init_db()
        self.load_index()

    def _connect(self) -> sqlite3.Connection:
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        return conn

    def _init_db(self) -> None:
        with self._connect() as conn:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS members (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    first_name TEXT,
                    last_name TEXT,
                    email TEXT,
                    membership_type TEXT,
                    membership_number TEXT,
                    includes_cart TEXT,
                    includes_range TEXT,
                    membership_amount_used INTEGER,
                    sheet_name TEXT,
                    row_number INTEGER,
                    membership_number_norm TEXT,
                    email_norm TEXT,
                    name_norm TEXT,
                    created_at TEXT,
                    updated_at TEXT
                )
                """
            )
            conn.execute("CREATE INDEX IF NOT EXISTS idx_members_number ON members (membership_number_norm)")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_members_email ON members (email_norm)")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_members_name ON members (name_norm)")

    def load_index(self) -> None:
        self._index = MemberIndex(self._fetch_all_records())

    def _fetch_all_records(self) -> list[MemberRecord]:
        with self._connect() as conn:
            rows = conn.execute(
                """
                SELECT id, first_name, last_name, email, membership_type, membership_number,
                       includes_cart, includes_range, membership_amount_used, sheet_name, row_number
                  FROM members
              ORDER BY last_name, first_name
                """
            ).fetchall()
        return [
            MemberRecord(
                record_id=row["id"],
                first_name=row["first_name"] or "",
                last_name=row["last_name"] or "",
                email=row["email"] or "",
                membership_type=row["membership_type"] or "",
                membership_number=row["membership_number"] or "",
                includes_cart=row["includes_cart"] or "No",
                includes_range=row["includes_range"] or "No",
                membership_amount_used=int(row["membership_amount_used"] or 0),
                sheet_name=row["sheet_name"],
                row_number=row["row_number"],
            )
            for row in rows
        ]

    def count_records(self) -> int:
        with self._connect() as conn:
            row = conn.execute("SELECT COUNT(*) AS total FROM members").fetchone()
            return int(row["total"] if row else 0)

    def lookup(self, scan_text: str) -> list[MemberRecord]:
        return self._index.lookup(scan_text)

    def sync_from_excel(self, file_obj) -> dict[str, int]:
        records = self._loader.load(file_obj)
        inserted, updated = self._sync_records(records)
        self.load_index()
        return {"total": len(records), "inserted": inserted, "updated": updated}

    def _sync_records(self, records: Iterable[MemberRecord]) -> tuple[int, int]:
        inserted = 0
        updated = 0
        timestamp = datetime.now().isoformat(timespec="seconds")
        with self._connect() as conn:
            for record in records:
                norms = self._normalized_fields(record)
                existing = self._find_existing(conn, norms)
                if existing:
                    if self._needs_update(existing, record, norms):
                        conn.execute(
                            """
                            UPDATE members
                               SET first_name = ?,
                                   last_name = ?,
                                   email = ?,
                                   membership_type = ?,
                                   membership_number = ?,
                                   includes_cart = ?,
                                   includes_range = ?,
                                   membership_amount_used = ?,
                                   sheet_name = ?,
                                   row_number = ?,
                                   membership_number_norm = ?,
                                   email_norm = ?,
                                   name_norm = ?,
                                   updated_at = ?
                             WHERE id = ?
                            """,
                            (
                                record.first_name,
                                record.last_name,
                                record.email,
                                record.membership_type,
                                record.membership_number,
                                record.includes_cart,
                                record.includes_range,
                                record.membership_amount_used,
                                record.sheet_name,
                                record.row_number,
                                norms["membership_number_norm"],
                                norms["email_norm"],
                                norms["name_norm"],
                                timestamp,
                                existing["id"],
                            ),
                        )
                        updated += 1
                else:
                    conn.execute(
                        """
                        INSERT INTO members (
                            first_name, last_name, email, membership_type, membership_number,
                            includes_cart, includes_range, membership_amount_used, sheet_name, row_number,
                            membership_number_norm, email_norm, name_norm, created_at, updated_at
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            record.first_name,
                            record.last_name,
                            record.email,
                            record.membership_type,
                            record.membership_number,
                            record.includes_cart,
                            record.includes_range,
                            record.membership_amount_used,
                            record.sheet_name,
                            record.row_number,
                            norms["membership_number_norm"],
                            norms["email_norm"],
                            norms["name_norm"],
                            timestamp,
                            timestamp,
                        ),
                    )
                    inserted += 1
        return inserted, updated

    def _normalized_fields(self, record: MemberRecord) -> dict[str, str]:
        return {
            "membership_number_norm": normalize(record.membership_number),
            "email_norm": normalize(record.email),
            "name_norm": normalize(f"{record.first_name} {record.last_name}"),
        }

    def _find_existing(self, conn: sqlite3.Connection, norms: dict[str, str]) -> sqlite3.Row | None:
        if norms["membership_number_norm"]:
            row = conn.execute(
                "SELECT * FROM members WHERE membership_number_norm = ? LIMIT 1",
                (norms["membership_number_norm"],),
            ).fetchone()
            if row:
                return row
        if norms["email_norm"]:
            row = conn.execute(
                "SELECT * FROM members WHERE email_norm = ? LIMIT 1",
                (norms["email_norm"],),
            ).fetchone()
            if row:
                return row
        if norms["name_norm"]:
            return conn.execute(
                "SELECT * FROM members WHERE name_norm = ? LIMIT 1",
                (norms["name_norm"],),
            ).fetchone()
        return None

    def _needs_update(self, row: sqlite3.Row, record: MemberRecord, norms: dict[str, str]) -> bool:
        return any(
            [
                (row["first_name"] or "") != record.first_name,
                (row["last_name"] or "") != record.last_name,
                (row["email"] or "") != record.email,
                (row["membership_type"] or "") != record.membership_type,
                (row["membership_number"] or "") != record.membership_number,
                (row["includes_cart"] or "") != record.includes_cart,
                (row["includes_range"] or "") != record.includes_range,
                int(row["membership_amount_used"] or 0) != int(record.membership_amount_used),
                (row["sheet_name"] or "") != (record.sheet_name or ""),
                int(row["row_number"] or 0) != int(record.row_number or 0),
                (row["membership_number_norm"] or "") != norms["membership_number_norm"],
                (row["email_norm"] or "") != norms["email_norm"],
                (row["name_norm"] or "") != norms["name_norm"],
            ]
        )
