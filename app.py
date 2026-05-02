from __future__ import annotations

import base64
import cgi
import csv
import hashlib
import html
import json
import os
import re
import shutil
import sqlite3
import subprocess
import sys
import secrets
from dataclasses import dataclass
from datetime import datetime
from http.cookies import SimpleCookie
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from io import BytesIO
from pathlib import Path
from urllib.parse import parse_qs, urlencode, urlparse

import openpyxl
from PIL import Image


APP_ICON_SOURCE = Path(__file__).resolve().parent / "Icon.png"
APP_VERSION = "1.0.0"


def inferred_icon_background(image: Image.Image) -> tuple[int, int, int, int]:
    rgba = image.convert("RGBA")
    w, h = rgba.size
    if w == 0 or h == 0:
        return (255, 255, 255, 255)
    samples = [rgba.getpixel((0, 0)), rgba.getpixel((w - 1, 0)), rgba.getpixel((0, h - 1)), rgba.getpixel((w - 1, h - 1))]
    luminance_values: list[float] = []
    for r, g, b, a in samples:
        if a < 20:
            continue
        luminance = (0.2126 * r) + (0.7152 * g) + (0.0722 * b)
        luminance_values.append(luminance)
    avg = (sum(luminance_values) / len(luminance_values)) if luminance_values else 255.0
    if avg < 128:
        return (0, 0, 0, 255)
    return (255, 255, 255, 255)


def build_icon_assets(source: Path, output_png: Path, output_ico: Path, target_size: int = 512) -> None:
    if not source.exists():
        return
    output_png.parent.mkdir(parents=True, exist_ok=True)
    image = Image.open(source).convert("RGBA")
    bg = inferred_icon_background(image)
    scale = target_size / max(1, image.height)
    new_w = max(1, int(round(image.width * scale)))
    resized = image.resize((new_w, target_size), Image.Resampling.LANCZOS)
    if resized.width > target_size:
        left = (resized.width - target_size) // 2
        resized = resized.crop((left, 0, left + target_size, target_size))
    canvas = Image.new("RGBA", (target_size, target_size), bg)
    x = (target_size - resized.width) // 2
    canvas.alpha_composite(resized, dest=(x, 0))
    canvas.save(output_png)
    canvas.save(output_ico, sizes=[(256, 256), (128, 128), (64, 64), (48, 48), (32, 32), (16, 16)])


def normalize(value: str) -> str:
    value = value or ""
    return re.sub(r"[^a-z0-9]", "", value.strip().lower())


def canonical_header(value: str) -> str:
    return normalize(value)


def split_first_name(first_name: str) -> list[str]:
    if not first_name:
        return []
    parts = [p.strip() for p in re.split(r"[/&]| and ", first_name, flags=re.IGNORECASE)]
    parts = [p for p in parts if p]
    return parts or [first_name.strip()]


def parse_yes_no(value: str) -> str:
    cleaned = (value or "").strip().lower()
    if cleaned in {"yes", "y", "true", "1"}:
        return "Yes"
    if cleaned in {"no", "n", "false", "0", ""}:
        return "No"
    return "No"


def safe_cell_text(value: str) -> str:
    text = (value or "").strip()
    if text.startswith(("=", "+", "-", "@")):
        raise ValueError("Values cannot start with =, +, -, or @.")
    return text


def safe_csv_value(value: str) -> str:
    text = str(value or "")
    if text.startswith(("=", "+", "-", "@")):
        return "'" + text
    return text


def layout_mode_for_width(width: int) -> str:
    return "compact" if width <= 880 else "wide"


def app_data_dir_for_platform(platform_name: str, home_path: Path) -> Path:
    if platform_name.startswith("win"):
        base = home_path / "AppData" / "Local"
    elif platform_name == "darwin":
        base = home_path / "Library" / "Application Support"
    else:
        base = home_path / ".local" / "share"
    return base / "MembershipVerifier"


class AuditTrail:
    def __init__(self, app_data_dir: Path) -> None:
        self.hidden_dir = app_data_dir / ".cache" / ".mvguard"
        self.audit_file = self.hidden_dir / ".act_journal.bin"
        self.seed_file = self.hidden_dir / ".seed.bin"
        self.state_file = self.hidden_dir / ".state.json"
        self.hidden_dir.mkdir(parents=True, exist_ok=True)
        self.seed = self._load_or_create_seed()
        self.last_hash = self._read_last_hash()

    def _load_or_create_seed(self) -> str:
        if self.seed_file.exists():
            return self.seed_file.read_text(encoding="utf-8").strip()
        seed = hashlib.sha256(os.urandom(64)).hexdigest()
        self.seed_file.write_text(seed, encoding="utf-8")
        return seed

    def _read_last_hash(self) -> str:
        if not self.audit_file.exists():
            return "0" * 64
        try:
            lines = self.audit_file.read_text(encoding="utf-8").splitlines()
            if not lines:
                return "0" * 64
            last = json.loads(lines[-1])
            return str(last.get("hash", "0" * 64))
        except Exception:
            return "0" * 64

    def verify_chain(self) -> bool:
        if not self.audit_file.exists():
            return True
        prev = "0" * 64
        try:
            for line in self.audit_file.read_text(encoding="utf-8").splitlines():
                if not line.strip():
                    continue
                item = json.loads(line)
                claimed_prev = item.get("prev_hash", "")
                claimed_hash = item.get("hash", "")
                payload = item.get("payload", {})
                base = json.dumps(payload, sort_keys=True, separators=(",", ":"))
                expected = hashlib.sha256(f"{prev}|{base}|{self.seed}".encode("utf-8")).hexdigest()
                if claimed_prev != prev or claimed_hash != expected:
                    return False
                prev = claimed_hash
            return True
        except Exception:
            return False

    def log(self, event: str, payload: dict[str, str | int | float | bool]) -> None:
        clean_payload = dict(payload)
        clean_payload["event"] = event
        clean_payload["ts"] = datetime.now().isoformat(timespec="seconds")
        base = json.dumps(clean_payload, sort_keys=True, separators=(",", ":"))
        digest = hashlib.sha256(f"{self.last_hash}|{base}|{self.seed}".encode("utf-8")).hexdigest()
        row = {"prev_hash": self.last_hash, "hash": digest, "payload": clean_payload}
        with self.audit_file.open("a", encoding="utf-8") as f:
            f.write(json.dumps(row, separators=(",", ":")) + "\n")
        self.last_hash = digest

    def get_state(self, key: str, default: str = "") -> str:
        if not self.state_file.exists():
            return default
        try:
            payload = json.loads(self.state_file.read_text(encoding="utf-8"))
            return str(payload.get(key, default))
        except Exception:
            return default

    def set_state(self, key: str, value: str) -> None:
        payload = {}
        if self.state_file.exists():
            try:
                payload = json.loads(self.state_file.read_text(encoding="utf-8"))
            except Exception:
                payload = {}
        payload[key] = value
        temp = self.state_file.with_suffix(".tmp")
        temp.write_text(json.dumps(payload, separators=(",", ":")), encoding="utf-8")
        os.replace(temp, self.state_file)


@dataclass
class MemberRecord:
    id: int
    first_name: str
    last_name: str
    email: str
    membership_type: str
    price_paid: str
    membership_number: str
    includes_cart: str
    includes_range: str
    membership_amount_used: int
    source_sheet: str
    source_row: int

    @property
    def signature(self) -> int:
        return self.id

    @property
    def display_name(self) -> str:
        return f"{self.first_name} {self.last_name}".strip()


@dataclass
class SheetConfig:
    sheet_name: str
    header_row: int
    index_map: dict[str, int]


@dataclass
class SyncResult:
    inserted: int = 0
    merged: int = 0
    scanned: int = 0


class MembershipDatabase:
    COL_ALIASES = {
        "first_name": {"firstname"},
        "last_name": {"lastname"},
        "email": {"email"},
        "membership_type": {"membershiptype"},
        "price_paid": {"pricepaid", "price", "paid"},
        "membership_number": {"membershipnumber", "membernumber"},
        "membership_amount_used": {"membershipamountused", "amountused", "membershipused"},
        "includes_cart": {"includescart", "incldescart", "includecart"},
        "includes_range": {"includesrange", "includerange"},
    }

    def __init__(self, db_path: Path) -> None:
        self.db_path = db_path
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        self._init_db()

    def _connect(self) -> sqlite3.Connection:
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA foreign_keys=ON")
        return conn

    def _init_db(self) -> None:
        with self._connect() as conn:
            conn.executescript(
                """
                CREATE TABLE IF NOT EXISTS members (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    member_key TEXT NOT NULL UNIQUE,
                    first_name TEXT NOT NULL DEFAULT '',
                    last_name TEXT NOT NULL DEFAULT '',
                    email TEXT NOT NULL DEFAULT '',
                    membership_type TEXT NOT NULL DEFAULT '',
                    price_paid TEXT NOT NULL DEFAULT '',
                    membership_number TEXT NOT NULL DEFAULT '',
                    includes_cart TEXT NOT NULL DEFAULT 'No',
                    includes_range TEXT NOT NULL DEFAULT 'No',
                    membership_amount_used INTEGER NOT NULL DEFAULT 0,
                    source_sheet TEXT NOT NULL DEFAULT '',
                    source_row INTEGER NOT NULL DEFAULT 0,
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL
                );

                CREATE INDEX IF NOT EXISTS idx_members_name ON members(last_name, first_name);
                CREATE INDEX IF NOT EXISTS idx_members_email ON members(email);
                CREATE INDEX IF NOT EXISTS idx_members_membership_number ON members(membership_number);
                CREATE INDEX IF NOT EXISTS idx_members_type ON members(membership_type);
                """
            )

    def _now(self) -> str:
        return datetime.now().isoformat(timespec="seconds")

    def _member_key(self, payload: dict[str, str | int]) -> str:
        membership_number = normalize(str(payload.get("membership_number", "")))
        email = normalize(str(payload.get("email", "")))
        first_name = normalize(str(payload.get("first_name", "")))
        last_name = normalize(str(payload.get("last_name", "")))
        membership_type = normalize(str(payload.get("membership_type", "")))
        price_paid = normalize(str(payload.get("price_paid", "")))
        if membership_number:
            return hashlib.sha256(f"mn|{membership_number}".encode("utf-8")).hexdigest()
        if email:
            return hashlib.sha256(f"em|{email}".encode("utf-8")).hexdigest()
        parts = [first_name, last_name, membership_type, price_paid]
        return hashlib.sha256(("name|" + "|".join(parts)).encode("utf-8")).hexdigest()

    def _record_from_row(self, row: sqlite3.Row) -> MemberRecord:
        return MemberRecord(
            id=int(row["id"]),
            first_name=str(row["first_name"]),
            last_name=str(row["last_name"]),
            email=str(row["email"]),
            membership_type=str(row["membership_type"]),
            price_paid=str(row["price_paid"]),
            membership_number=str(row["membership_number"]),
            includes_cart=str(row["includes_cart"]),
            includes_range=str(row["includes_range"]),
            membership_amount_used=int(row["membership_amount_used"]),
            source_sheet=str(row["source_sheet"]),
            source_row=int(row["source_row"]),
        )

    def all_records(self) -> list[MemberRecord]:
        with self._connect() as conn:
            rows = conn.execute(
                "SELECT * FROM members ORDER BY last_name COLLATE NOCASE, first_name COLLATE NOCASE, id"
            ).fetchall()
        return [self._record_from_row(row) for row in rows]

    def get_record(self, member_id: int) -> MemberRecord | None:
        with self._connect() as conn:
            row = conn.execute("SELECT * FROM members WHERE id = ?", (member_id,)).fetchone()
        return self._record_from_row(row) if row else None

    def update_record(self, member_id: int, updates: dict[str, str | int]) -> MemberRecord:
        record = self.get_record(member_id)
        if not record:
            raise ValueError("Selected member record was not found.")

        fields: dict[str, str | int] = {}
        if "first_name" in updates:
            fields["first_name"] = safe_cell_text(str(updates["first_name"]))
        if "last_name" in updates:
            fields["last_name"] = safe_cell_text(str(updates["last_name"]))
        if "email" in updates:
            fields["email"] = safe_cell_text(str(updates["email"]))
        if "membership_type" in updates:
            fields["membership_type"] = safe_cell_text(str(updates["membership_type"]))
        if "price_paid" in updates:
            fields["price_paid"] = safe_cell_text(str(updates["price_paid"]))
        if "membership_number" in updates:
            fields["membership_number"] = safe_cell_text(str(updates["membership_number"]))
        if "includes_cart" in updates:
            fields["includes_cart"] = parse_yes_no(str(updates["includes_cart"]))
        if "includes_range" in updates:
            fields["includes_range"] = parse_yes_no(str(updates["includes_range"]))
        if "membership_amount_used" in updates:
            try:
                amount = int(str(updates["membership_amount_used"]).strip())
            except ValueError as exc:
                raise ValueError("Membership Amount Used must be a whole number.") from exc
            if amount < 0:
                raise ValueError("Membership Amount Used cannot be negative.")
            fields["membership_amount_used"] = amount

        if not fields:
            return record

        key_payload = {
            "first_name": fields.get("first_name", record.first_name),
            "last_name": fields.get("last_name", record.last_name),
            "email": fields.get("email", record.email),
            "membership_type": fields.get("membership_type", record.membership_type),
            "price_paid": fields.get("price_paid", record.price_paid),
            "membership_number": fields.get("membership_number", record.membership_number),
            "includes_cart": fields.get("includes_cart", record.includes_cart),
            "includes_range": fields.get("includes_range", record.includes_range),
        }
        new_key = self._member_key(key_payload)

        with self._connect() as conn:
            conn.execute(
                f"UPDATE members SET {', '.join(f'{k} = ?' for k in fields)} , member_key = ?, updated_at = ? WHERE id = ?",
                (*fields.values(), new_key, self._now(), member_id),
            )
            row = conn.execute("SELECT * FROM members WHERE id = ?", (member_id,)).fetchone()
        if not row:
            raise RuntimeError("Member update failed.")
        return self._record_from_row(row)

    def upsert_new_member(self, payload: dict[str, str | int]) -> tuple[str, MemberRecord]:
        return self.upsert_source_row(payload)

    def increment_usage(self, member_id: int, delta: int) -> MemberRecord:
        record = self.get_record(member_id)
        if not record:
            raise ValueError("Member not found.")
        new_value = max(0, record.membership_amount_used + delta)
        with self._connect() as conn:
            conn.execute(
                "UPDATE members SET membership_amount_used = ?, updated_at = ? WHERE id = ?",
                (new_value, self._now(), member_id),
            )
        updated = self.get_record(member_id)
        if not updated:
            raise RuntimeError("Could not reload member after usage update.")
        return updated

    def upsert_source_row(self, payload: dict[str, str | int]) -> tuple[str, MemberRecord]:
        now = self._now()
        member_key = self._member_key(payload)
        payload = dict(payload)
        payload["member_key"] = member_key
        payload.setdefault("created_at", now)
        payload["updated_at"] = now

        with self._connect() as conn:
            row = conn.execute("SELECT * FROM members WHERE member_key = ?", (member_key,)).fetchone()
            if not row:
                conn.execute(
                    """
                    INSERT INTO members (
                        member_key, first_name, last_name, email, membership_type, price_paid,
                        membership_number, includes_cart, includes_range, membership_amount_used,
                        source_sheet, source_row, created_at, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        member_key,
                        str(payload.get("first_name", "")),
                        str(payload.get("last_name", "")),
                        str(payload.get("email", "")),
                        str(payload.get("membership_type", "")),
                        str(payload.get("price_paid", "")),
                        str(payload.get("membership_number", "")),
                        parse_yes_no(str(payload.get("includes_cart", "No"))),
                        parse_yes_no(str(payload.get("includes_range", "No"))),
                        int(payload.get("membership_amount_used", 0) or 0),
                        str(payload.get("source_sheet", "")),
                        int(payload.get("source_row", 0) or 0),
                        now,
                        now,
                    ),
                )
                inserted = conn.execute("SELECT * FROM members WHERE member_key = ?", (member_key,)).fetchone()
                if not inserted:
                    raise RuntimeError("Insert failed.")
                return "inserted", self._record_from_row(inserted)

            merge_fields: dict[str, str | int] = {}
            for field in [
                "first_name",
                "last_name",
                "email",
                "membership_type",
                "price_paid",
                "membership_number",
                "includes_cart",
                "includes_range",
                "source_sheet",
            ]:
                incoming = str(payload.get(field, "")).strip()
                existing = str(row[field]).strip()
                if not existing and incoming:
                    merge_fields[field] = parse_yes_no(incoming) if field in {"includes_cart", "includes_range"} else incoming

            incoming_amount = int(payload.get("membership_amount_used", 0) or 0)
            if int(row["membership_amount_used"]) == 0 and incoming_amount > 0:
                merge_fields["membership_amount_used"] = incoming_amount

            merge_fields["source_row"] = int(payload.get("source_row", row["source_row"]) or row["source_row"])
            merge_fields["updated_at"] = now

            if merge_fields:
                assignments = ", ".join(f"{field} = ?" for field in merge_fields)
                conn.execute(
                    f"UPDATE members SET {assignments} WHERE member_key = ?",
                    (*merge_fields.values(), member_key),
                )
                row = conn.execute("SELECT * FROM members WHERE member_key = ?", (member_key,)).fetchone()
                if not row:
                    raise RuntimeError("Merge failed.")
                return "merged", self._record_from_row(row)

            return "unchanged", self._record_from_row(row)

    def _find_sheet_config(self, sheet_name: str, sheet) -> SheetConfig | None:
        header_row = 0
        raw_headers: list[str] = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=30, values_only=True), start=1):
            cells = [str(c).strip() if c is not None else "" for c in row]
            canon = {canonical_header(c) for c in cells if c}
            if "firstname" in canon and "lastname" in canon:
                header_row = row_idx
                raw_headers = cells
                break
        if not header_row:
            return None

        source = {canonical_header(v): idx for idx, v in enumerate(raw_headers, start=1) if v}
        index_map: dict[str, int] = {}
        for logical, aliases in self.COL_ALIASES.items():
            for alias in aliases:
                if alias in source:
                    index_map[logical] = source[alias]
                    break

        required = {"first_name", "last_name", "membership_type"}
        if not required.issubset(index_map.keys()):
            return None
        return SheetConfig(sheet_name=sheet_name, header_row=header_row, index_map=index_map)

    def _cell_text(self, row, idx: int | None) -> str:
        if not idx:
            return ""
        value = row[idx - 1]
        if value is None:
            return ""
        return str(value).strip()

    def _cell_int(self, row, idx: int | None) -> int:
        if not idx:
            return 0
        value = row[idx - 1]
        if value is None or str(value).strip() == "":
            return 0
        try:
            return int(float(str(value).strip()))
        except ValueError:
            return 0

    def sync_from_workbook(self, excel_path: Path) -> SyncResult:
        workbook = openpyxl.load_workbook(excel_path, data_only=True)
        result = SyncResult()
        try:
            for sheet_name in workbook.sheetnames:
                if "total" in sheet_name.lower():
                    continue
                sheet = workbook[sheet_name]
                config = self._find_sheet_config(sheet_name, sheet)
                if not config:
                    continue
                for row_idx, row in enumerate(
                    sheet.iter_rows(min_row=config.header_row + 1, max_row=sheet.max_row, values_only=True),
                    start=config.header_row + 1,
                ):
                    first_name = self._cell_text(row, config.index_map.get("first_name"))
                    last_name = self._cell_text(row, config.index_map.get("last_name"))
                    email = self._cell_text(row, config.index_map.get("email"))
                    membership_number = self._cell_text(row, config.index_map.get("membership_number"))
                    membership_type = self._cell_text(row, config.index_map.get("membership_type"))
                    price_paid = self._cell_text(row, config.index_map.get("price_paid"))
                    if not (first_name or last_name or email or membership_number):
                        continue
                    result.scanned += 1
                    status, _ = self.upsert_source_row(
                        {
                            "first_name": first_name,
                            "last_name": last_name,
                            "email": email,
                            "membership_type": membership_type,
                            "price_paid": price_paid,
                            "membership_number": membership_number,
                            "includes_cart": self._cell_text(row, config.index_map.get("includes_cart")) or "No",
                            "includes_range": self._cell_text(row, config.index_map.get("includes_range")) or "No",
                            "membership_amount_used": self._cell_int(row, config.index_map.get("membership_amount_used")),
                            "source_sheet": sheet_name,
                            "source_row": row_idx,
                        }
                    )
                    if status == "inserted":
                        result.inserted += 1
                    elif status == "merged":
                        result.merged += 1
        finally:
            workbook.close()
        return result

    def lookup(self, scan_text: str) -> list[MemberRecord]:
        raw = (scan_text or "").strip()
        if not raw:
            return []
        records = self.all_records()
        key = normalize(raw)
        matches: list[MemberRecord] = []
        if key:
            matches.extend([r for r in records if normalize(r.membership_number) == key and r.membership_number])
        if not matches and key:
            matches.extend([r for r in records if normalize(r.email) == key and r.email])
        if not matches:
            for candidate in self._name_candidates(raw):
                ckey = normalize(candidate)
                if not ckey:
                    continue
                candidate_matches = [r for r in records if ckey in {normalize(r.display_name), normalize(f"{r.last_name} {r.first_name}"), normalize(f"{r.first_name} {r.last_name}")}]
                matches.extend(candidate_matches)
        if not matches:
            needle = normalize(raw)
            for record in records:
                haystack = normalize(f"{record.first_name} {record.last_name} {record.email} {record.membership_number}")
                if needle and needle in haystack:
                    matches.append(record)
        unique: list[MemberRecord] = []
        seen = set()
        for item in matches:
            if item.id in seen:
                continue
            seen.add(item.id)
            unique.append(item)
        return unique

    def _name_candidates(self, raw: str) -> list[str]:
        values = [raw]
        if "," in raw:
            parts = [p.strip() for p in raw.split(",") if p.strip()]
            if len(parts) >= 2:
                values.extend([f"{parts[1]} {parts[0]}", f"{parts[0]} {parts[1]}"])
        else:
            parts = [p for p in raw.split() if p]
            if len(parts) >= 2:
                values.extend([f"{parts[0]} {parts[-1]}", f"{parts[-1]} {parts[0]}"])
        return values

    def stats(self) -> dict[str, int]:
        with self._connect() as conn:
            total = conn.execute("SELECT COUNT(*) FROM members").fetchone()[0]
            used = conn.execute("SELECT COALESCE(SUM(membership_amount_used), 0) FROM members").fetchone()[0]
        return {"members": int(total), "usage": int(used)}


class MembershipWebApp:
    def __init__(self, db: MembershipDatabase, audit: AuditTrail, base_url: str = "") -> None:
        self.db = db
        self.audit = audit
        self.base_url = base_url.rstrip("/")
        self.last_sync: SyncResult | None = None
        self.last_sync_name: str = ""

    def html_page(self, title: str, body: str) -> bytes:
        page = f"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{html.escape(title)}</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
<style>
*:{{box-sizing:border-box;margin:0;padding:0}}
:root{{
  --primary:#2563eb;
  --primary-hover:#1d4ed8;
  --primary-light:#eff6ff;
  --success:#10b981;
  --warning:#f59e0b;
  --danger:#ef4444;
  --gray-50:#f9fafb;
  --gray-100:#f3f4f6;
  --gray-200:#e5e7eb;
  --gray-300:#d1d5db;
  --gray-400:#9ca3af;
  --gray-500:#6b7280;
  --gray-600:#4b5563;
  --gray-700:#374151;
  --gray-800:#1f2937;
  --gray-900:#111827;
  --shadow-sm:0 1px 2px 0 rgb(0 0 0 / 0.05);
  --shadow:0 1px 3px 0 rgb(0 0 0 / 0.1),0 1px 2px -1px rgb(0 0 0 / 0.1);
  --shadow-md:0 4px 6px -1px rgb(0 0 0 / 0.1),0 2px 4px -2px rgb(0 0 0 / 0.1);
  --radius:12px;
  --radius-sm:8px;
}}
body{{font-family:'Inter',system-ui,-apple-system,'Segoe UI',Roboto,sans-serif;background:linear-gradient(135deg,#f8fafc 0%,#e2e8f0 100%);color:var(--gray-800);min-height:100vh;line-height:1.5}}
.container{{max-width:1280px;margin:0 auto;padding:0 24px}}
header{{background:linear-gradient(135deg,#1e293b 0%,#0f172a 100%);padding:20px 0;color:#fff;box-shadow:var(--shadow-md);position:relative;overflow:hidden}}
header::before{{content:'';position:absolute;top:0;right:0;width:300px;height:100%;background:linear-gradient(90deg,transparent,rgba(255,255,255,0.05));pointer-events:none}}
header .container{{display:flex;align-items:center;justify-content:space-between}}
.logo{{display:flex;align-items:center;gap:12px}}
.logo-icon{{width:40px;height:40px;background:linear-gradient(135deg,#3b82f6,#1d4ed8);border-radius:var(--radius-sm);display:flex;align-items:center;justify-content:center;font-weight:700;font-size:18px}}
.logo-text{{font-size:22px;font-weight:700;letter-spacing:-0.5px}}
.logo-version{{font-size:12px;background:rgba(255,255,255,0.15);padding:4px 10px;border-radius:20px;font-weight:500}}
main{{padding:32px 0}}
.hero{{background:#fff;border-radius:var(--radius);padding:32px;box-shadow:var(--shadow-md);margin-bottom:24px;position:relative;overflow:hidden}}
.hero::before{{content:'';position:absolute;top:0;left:0;right:0;height:4px;background:linear-gradient(90deg,#3b82f6,#8b5cf6,#06b6d4)}}
.hero-content{{display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:20px}}
.hero-title{{font-size:28px;font-weight:700;color:var(--gray-900);margin-bottom:8px}}
.hero-subtitle{{color:var(--gray-500);font-size:15px;max-width:600px}}
.hero-actions{{display:flex;gap:12px}}
.btn{{display:inline-flex;align-items:center;gap:8px;padding:12px 20px;border-radius:var(--radius-sm);font-weight:600;font-size:14px;text-decoration:none;transition:all 0.2s;border:none;cursor:pointer}}
.btn-primary{{background:var(--primary);color:#fff}}
.btn-primary:hover{{background:var(--primary-hover);transform:translateY(-1px);box-shadow:var(--shadow-md)}}
.btn-secondary{{background:var(--gray-100);color:var(--gray-700)}}
.btn-secondary:hover{{background:var(--gray-200)}}
.btn-sm{{padding:8px 14px;font-size:13px}}
.stats-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(200px,1fr));gap:16px;margin-bottom:24px}}
.stat-card{{background:#fff;border-radius:var(--radius);padding:20px;box-shadow:var(--shadow-sm);border:1px solid var(--gray-100);transition:transform 0.2s,box-shadow 0.2s}}
.stat-card:hover{{transform:translateY(-2px);box-shadow:var(--shadow-md)}}
.stat-label{{color:var(--gray-500);font-size:13px;font-weight:500;text-transform:uppercase;letter-spacing:0.5px;margin-bottom:4px}}
.stat-value{{font-size:32px;font-weight:700;color:var(--gray-900)}}
.stat-value.success{{color:var(--success)}}
.grid{{display:grid;grid-template-columns:repeat(2,1fr);gap:24px}}
@media (max-width:900px){{.grid{{grid-template-columns:1fr}}}}
.card{{background:#fff;border-radius:var(--radius);padding:24px;box-shadow:var(--shadow-sm);border:1px solid var(--gray-100)}}
.card-header{{display:flex;align-items:center;justify-content:space-between;margin-bottom:16px}}
.card-title{{font-size:18px;font-weight:600;color:var(--gray-900)}}
.card-description{{color:var(--gray-500);font-size:14px;margin-bottom:16px;line-height:1.6}}
.full{{grid-column:1/-1}}
.form-group{{margin-bottom:16px}}
.form-label{{display:block;font-size:14px;font-weight:500;color:var(--gray-700);margin-bottom:6px}}
.form-input,.form-select{{width:100%;padding:12px 14px;border:1px solid var(--gray-200);border-radius:var(--radius-sm);font-size:14px;transition:border-color 0.2s,box-shadow 0.2s;background:#fff}}
.form-input:focus,.form-select:focus{{outline:none;border-color:var(--primary);box-shadow:0 0 0 3px rgba(37,99,235,0.1)}}
.form-input::placeholder{{color:var(--gray-400)}}
.form-row{{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:16px}}
.form-actions{{display:flex;gap:12px;margin-top:20px}}
.member-card{{background:#fff;border-radius:var(--radius);padding:20px;box-shadow:var(--shadow-sm);border:1px solid var(--gray-100);margin-bottom:16px;transition:all 0.2s}}
.member-card:hover{{border-color:var(--primary);box-shadow:var(--shadow-md)}}
.member-card.selected{{border-color:var(--primary);background:var(--primary-light)}}
.member-name{{font-size:18px;font-weight:600;color:var(--gray-900);margin-bottom:4px}}
.member-meta{{color:var(--gray-500);font-size:14px}}
.member-details{{display:flex;flex-wrap:wrap;gap:16px;margin:12px 0;padding:12px;background:var(--gray-50);border-radius:var(--radius-sm);font-size:13px}}
.member-detail{{display:flex;align-items:center;gap:6px}}
.member-detail-label{{color:var(--gray-500)}}
.member-detail-value{{font-weight:500;color:var(--gray-700)}}
.member-actions{{display:flex;gap:8px;flex-wrap:wrap;margin-top:16px;padding-top:16px;border-top:1px solid var(--gray-100)}}
.member-edit-form{{margin-top:16px;padding-top:16px;border-top:1px solid var(--gray-200)}}
.msg{{padding:16px 20px;border-radius:var(--radius-sm);margin-bottom:20px;font-size:14px}}
.msg-success{{background:#ecfdf5;color:#065f46;border:1px solid #a7f3d0}}
.msg-error{{background:#fef2f2;color:#991b1b;border:1px solid #fecaca}}
.msg-info{{background:var(--primary-light);color:#1e40af;border:1px solid:#bfdbfe}}
.empty-state{{text-align:center;padding:48px 24px;color:var(--gray-500)}}
.empty-state-icon{{font-size:48px;margin-bottom:16px;opacity:0.5}}
.empty-state-text{{font-size:16px}}
.section{{margin-bottom:32px}}
.section-title{{font-size:20px;font-weight:600;color:var(--gray-900);margin-bottom:16px}}
.file-drop{{border:2px dashed var(--gray-300);border-radius:var(--radius);padding:32px;text-align:center;cursor:pointer;transition:all 0.2s;background:var(--gray-50)}}
.file-drop:hover{{border-color:var(--primary);background:var(--primary-light)}}
.file-drop-input{{display:none}}
</style>
</head>
<body>
<header>
<div class="container">
<div class="logo">
<div class="logo-icon">M</div>
<span class="logo-text">Membership Manager</span>
</div>
<span class="logo-version">v{APP_VERSION}</span>
</div>
</header>
<main>
<div class="container">{body}</div>
</main>
</body>
</html>"""
        return page.encode("utf-8")

    def render_form_token(self, token: str) -> str:
        return f"<input type='hidden' name='csrf_token' value='{html.escape(token)}'>"

    def render_dashboard(self, *, q: str = "", member_id: int | None = None, message: str = "", level: str = "", csrf_token: str = "") -> bytes:
        stats = self.db.stats()
        records = self.db.lookup(q) if q else []
        selected = self.db.get_record(member_id) if member_id else None
        sync_html = ""
        if self.last_sync:
            sync_html = f"<div class='pill'>Last sync: +{self.last_sync.inserted} inserted, {self.last_sync.merged} merged from {html.escape(self.last_sync_name or 'upload')}</div>"
        message_html = f"<div class='msg'>{html.escape(message)}</div>" if message else ""
        cards = "".join(
            self.render_member_card(record, selected_id=selected.id if selected else None, csrf_token=csrf_token) for record in records
        ) or """<div class='empty-state'>
<div class='empty-state-icon'>📋</div>
<div class='empty-state-text'>Search for a member or import an Excel file to get started</div>
</div>"""
        msg_class = "msg-success" if "error" not in message.lower() and "failed" not in message.lower() else "msg-error"
        body = f"""
<div class='hero'>
<div class='hero-content'>
<div>
<div class='hero-title'>Membership Manager</div>
<div class='hero-subtitle'>Import Excel files, search members, track usage, and manage your database from one modern interface.</div>
</div>
<div class='hero-actions'>
<a href='/export/sqlite' class='btn btn-primary'>📥 Export Database</a>
<a href='/?message=Ready for your next scan' class='btn btn-secondary'>🔄 Refresh</a>
</div>
</div>
</div>
{message_html}
<div class='stats-grid'>
<div class='stat-card'>
<div class='stat-label'>Total Members</div>
<div class='stat-value'>{stats['members']}</div>
</div>
<div class='stat-card'>
<div class='stat-label'>Total Usage</div>
<div class='stat-value success'>{stats['usage']}</div>
</div>
{sync_html}
</div>
<div class='grid'>
<section class='card'>
<div class='card-header'>
<div class='card-title'>📊 Sync Excel</div>
</div>
<p class='card-description'>Upload an Excel workbook to import or update members. New records are inserted, existing records are merged.</p>
<form method='post' action='/sync' enctype='multipart/form-data'>
{self.render_form_token(csrf_token)}
<div class='form-group'>
<label class='form-label'>Select Workbook</label>
<input type='file' name='workbook' accept='.xlsx,.xlsm' class='form-input' required>
</div>
<button type='submit' class='btn btn-primary'>Import Members</button>
</form>
</section>
<section class='card'>
<div class='card-header'>
<div class='card-title'>🔍 Search / Scan</div>
</div>
<p class='card-description'>Search by membership number, email, or member name.</p>
<form method='get' action='/'>
<div class='form-group'>
<label class='form-label'>Search Value</label>
<input name='q' value='{html.escape(q)}' class='form-input' placeholder='Enter membership number, email, or name'>
</div>
<button type='submit' class='btn btn-primary'>Search</button>
</form>
</section>
<section class='card full'>
<div class='card-header'>
<div class='card-title'>➕ Add New Member</div>
</div>
<form method='post' action='/members/create'>
{self.render_form_token(csrf_token)}
<div class='form-row'>
<div class='form-group'><label class='form-label'>First Name</label><input name='first_name' class='form-input' value=''></div>
<div class='form-group'><label class='form-label'>Last Name</label><input name='last_name' class='form-input' value=''></div>
<div class='form-group'><label class='form-label'>Email</label><input name='email' class='form-input' value='' type='email'></div>
<div class='form-group'><label class='form-label'>Membership Type</label><input name='membership_type' class='form-input' value=''></div>
<div class='form-group'><label class='form-label'>Price Paid</label><input name='price_paid' class='form-input' value=''></div>
<div class='form-group'><label class='form-label'>Membership Number</label><input name='membership_number' class='form-input' value=''></div>
<div class='form-group'><label class='form-label'>Includes Cart</label><select name='includes_cart' class='form-select'><option selected>No</option><option>Yes</option></select></div>
<div class='form-group'><label class='form-label'>Includes Range</label><select name='includes_range' class='form-select'><option selected>No</option><option>Yes</option></select></div>
</div>
<div class='form-actions'><button type='submit' class='btn btn-primary'>Create Member</button></div>
</form>
</section>
<section class='card full'>
<div class='card-header'>
<div class='card-title'>📋 Results</div>
</div>
{cards}
</section>
</div>
"""
        return self.html_page("Membership WebApp", body)

    def render_member_card(self, record: MemberRecord, selected_id: int | None = None, csrf_token: str = "") -> str:
        selected = " selected" if selected_id == record.id else ""
        return f"""
<div class='member-card{selected}'>
<div class='member-name'>{html.escape(record.display_name or 'Unnamed member')} <span style='color:var(--gray-400);font-weight:400;font-size:14px'>#{html.escape(record.membership_number or 'N/A')}</span></div>
<div class='member-meta'>{html.escape(record.membership_type or 'Unknown Type')}</div>
<div class='member-details'>
<div class='member-detail'><span class='member-detail-label'>📧</span><span class='member-detail-value'>{html.escape(record.email or 'No email')}</span></div>
<div class='member-detail'><span class='member-detail-label'>🛒</span><span class='member-detail-value'>Cart: {html.escape(parse_yes_no(record.includes_cart))}</span></div>
<div class='member-detail'><span class='member-detail-label'>🎯</span><span class='member-detail-value'>Range: {html.escape(parse_yes_no(record.includes_range))}</span></div>
<div class='member-detail'><span class='member-detail-label'>📊</span><span class='member-detail-value'>Used: {record.membership_amount_used}</span></div>
</div>
<div class='member-actions'>
<form method='get' action='/'><input type='hidden' name='member' value='{record.id}'><input type='hidden' name='q' value='{html.escape(record.display_name)}'><button type='submit' class='btn btn-secondary btn-sm'>✏️ Edit</button></form>
<form method='post' action='/members/{record.id}/usage'>{self.render_form_token(csrf_token)}<input type='hidden' name='delta' value='1'><button type='submit' class='btn btn-primary btn-sm'>➕ Usage</button></form>
<form method='post' action='/members/{record.id}/usage'>{self.render_form_token(csrf_token)}<input type='hidden' name='delta' value='-1'><button type='submit' class='btn btn-secondary btn-sm'>➖ Usage</button></form>
</div>
<div class='member-edit-form'>
<form method='post' action='/members/{record.id}/update'>
{self.render_form_token(csrf_token)}
<div class='form-row'>
<div class='form-group'><label class='form-label'>First Name</label><input name='first_name' class='form-input' value='{html.escape(record.first_name)}'></div>
<div class='form-group'><label class='form-label'>Last Name</label><input name='last_name' class='form-input' value='{html.escape(record.last_name)}'></div>
<div class='form-group'><label class='form-label'>Email</label><input name='email' class='form-input' value='{html.escape(record.email)}'></div>
<div class='form-group'><label class='form-label'>Membership Type</label><input name='membership_type' class='form-input' value='{html.escape(record.membership_type)}'></div>
<div class='form-group'><label class='form-label'>Price Paid</label><input name='price_paid' class='form-input' value='{html.escape(record.price_paid)}'></div>
<div class='form-group'><label class='form-label'>Membership Number</label><input name='membership_number' class='form-input' value='{html.escape(record.membership_number)}'></div>
<div class='form-group'><label class='form-label'>Includes Cart</label><select name='includes_cart' class='form-select'><option {'selected' if parse_yes_no(record.includes_cart)=='Yes' else ''}>Yes</option><option {'selected' if parse_yes_no(record.includes_cart)=='No' else ''}>No</option></select></div>
<div class='form-group'><label class='form-label'>Includes Range</label><select name='includes_range' class='form-select'><option {'selected' if parse_yes_no(record.includes_range)=='Yes' else ''}>Yes</option><option {'selected' if parse_yes_no(record.includes_range)=='No' else ''}>No</option></select></div>
<div class='form-group'><label class='form-label'>Amount Used</label><input name='membership_amount_used' class='form-input' value='{record.membership_amount_used}' type='number'></div>
</div>
<div class='form-actions'><button type='submit' class='btn btn-primary btn-sm'>💾 Save Changes</button></div>
</form>
</div>
</div>
"""

    def handle_sync(self, uploaded: tuple[str, bytes]) -> tuple[int, str]:
        filename, data = uploaded
        suffix = Path(filename or "upload.xlsx").suffix.lower()
        if suffix not in {".xlsx", ".xlsm"}:
            return HTTPStatus.BAD_REQUEST, "Please upload an .xlsx or .xlsm workbook."
        temp_dir = Path(tempfile_dir())
        temp_dir.mkdir(parents=True, exist_ok=True)
        temp_path = temp_dir / f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{Path(filename).name}"
        temp_path.write_bytes(data)
        try:
            sync_result = self.db.sync_from_workbook(temp_path)
            self.last_sync = sync_result
            self.last_sync_name = filename
            self.audit.log(
                "excel_synced",
                {"filename": filename, "inserted": sync_result.inserted, "merged": sync_result.merged, "scanned": sync_result.scanned},
            )
            return HTTPStatus.SEE_OTHER, "/?message=" + urlencode({"message": f"Imported {sync_result.scanned} rows from {filename}. Inserted {sync_result.inserted}, merged {sync_result.merged}."})[8:]
        finally:
            temp_path.unlink(missing_ok=True)

    def handle(self, handler: BaseHTTPRequestHandler) -> None:
        parsed = urlparse(handler.path)
        params = parse_qs(parsed.query)
        message = params.get("message", [""])[0]
        q = params.get("q", [""])[0]
        member = params.get("member", [""])[0]
        member_id = int(member) if member.isdigit() else None

        if parsed.path in {"/", "/index.html"} and handler.command == "GET":
            csrf_token = handler.get_csrf_token()
            handler.respond_bytes(self.render_dashboard(q=q, member_id=member_id, message=message, csrf_token=csrf_token))
            return
        try:
            if parsed.path == "/sync" and handler.command == "POST":
                status, target = self._handle_sync_request(handler)
                handler.send_response(status)
                if status == HTTPStatus.SEE_OTHER:
                    handler.send_header("Location", target)
                    handler.end_headers()
                    return
                handler.respond_text(str(target), status=status)
                return
            if parsed.path == "/members/create" and handler.command == "POST":
                self._handle_create_request(handler)
                return
            if parsed.path.endswith("/update") and handler.command == "POST":
                self._handle_update_request(handler, parsed.path)
                return
            if parsed.path.endswith("/usage") and handler.command == "POST":
                self._handle_usage_request(handler, parsed.path)
                return
            if parsed.path == "/export/sqlite" and handler.command == "GET":
                self._handle_export_request(handler)
                return
        except PermissionError as exc:
            handler.respond_text(str(exc), status=HTTPStatus.FORBIDDEN)
            return
        except Exception as exc:
            handler.respond_text(f"Request failed: {exc}", status=HTTPStatus.BAD_REQUEST)
            return
        handler.respond_text("Not found", status=HTTPStatus.NOT_FOUND)

    def _handle_sync_request(self, handler: AuthenticatedHandler) -> tuple[int, str]:
        form = parse_multipart_form(handler)
        self._validate_csrf(handler, form)
        upload = form.get("workbook")
        if not upload:
            return HTTPStatus.BAD_REQUEST, "Missing workbook upload."
        filename, data = upload
        return self.handle_sync((filename, data))

    def _handle_update_request(self, handler: AuthenticatedHandler, path: str) -> None:
        member_id = extract_member_id(path)
        if member_id is None:
            handler.respond_text("Invalid member id", status=HTTPStatus.BAD_REQUEST)
            return
        form = parse_multipart_form(handler)
        self._validate_csrf(handler, form)
        try:
            updated = self.db.update_record(
                member_id,
                {
                    "first_name": form.get("first_name", ""),
                    "last_name": form.get("last_name", ""),
                    "email": form.get("email", ""),
                    "membership_type": form.get("membership_type", ""),
                    "price_paid": form.get("price_paid", ""),
                    "membership_number": form.get("membership_number", ""),
                    "includes_cart": form.get("includes_cart", "No"),
                    "includes_range": form.get("includes_range", "No"),
                    "membership_amount_used": form.get("membership_amount_used", "0"),
                },
            )
            self.audit.log("member_updated", {"member_id": updated.id, "membership_number": updated.membership_number})
            handler.send_response(HTTPStatus.SEE_OTHER)
            handler.send_header("Location", f"/?member={updated.id}&message={urlencode({'message': 'Member saved.'})[8:]}")
            handler.end_headers()
        except Exception as exc:
            handler.respond_text(f"Save error: {exc}", status=HTTPStatus.BAD_REQUEST)

    def _handle_usage_request(self, handler: AuthenticatedHandler, path: str) -> None:
        member_id = extract_member_id(path)
        if member_id is None:
            handler.respond_text("Invalid member id", status=HTTPStatus.BAD_REQUEST)
            return
        form = parse_multipart_form(handler)
        self._validate_csrf(handler, form)
        try:
            delta = int(str(form.get("delta", "1")).strip())
        except ValueError:
            delta = 1
        updated = self.db.increment_usage(member_id, delta)
        self.audit.log("usage_updated", {"member_id": updated.id, "delta": delta, "value": updated.membership_amount_used})
        handler.send_response(HTTPStatus.SEE_OTHER)
        handler.send_header("Location", f"/?member={updated.id}&message={urlencode({'message': 'Usage updated.'})[8:]}")
        handler.end_headers()

    def _handle_create_request(self, handler: AuthenticatedHandler) -> None:
        form = parse_multipart_form(handler)
        self._validate_csrf(handler, form)
        try:
            status, record = self.db.upsert_new_member(
                {
                    "first_name": form.get("first_name", ""),
                    "last_name": form.get("last_name", ""),
                    "email": form.get("email", ""),
                    "membership_type": form.get("membership_type", ""),
                    "price_paid": form.get("price_paid", ""),
                    "membership_number": form.get("membership_number", ""),
                    "includes_cart": form.get("includes_cart", "No"),
                    "includes_range": form.get("includes_range", "No"),
                    "membership_amount_used": form.get("membership_amount_used", "0"),
                    "source_sheet": "manual",
                    "source_row": 0,
                }
            )
            self.audit.log("member_created", {"member_id": record.id, "membership_number": record.membership_number})
            handler.send_response(HTTPStatus.SEE_OTHER)
            handler.send_header("Location", f"/?member={record.id}&message={urlencode({'message': 'Member created.'})[8:]}")
            handler.end_headers()
        except Exception as exc:
            handler.respond_text(f"Create error: {exc}", status=HTTPStatus.BAD_REQUEST)

    def _handle_export_request(self, handler: BaseHTTPRequestHandler) -> None:
        path = self.db.db_path
        if not path.exists():
            handler.respond_text("Database not found.", status=HTTPStatus.NOT_FOUND)
            return
        data = path.read_bytes()
        handler.send_response(HTTPStatus.OK)
        handler._secure_headers()
        handler.send_header("Content-Type", "application/x-sqlite3")
        handler.send_header("Content-Disposition", 'attachment; filename="members.sqlite3"')
        handler.send_header("Content-Length", str(len(data)))
        handler.end_headers()
        handler.wfile.write(data)

    def _validate_csrf(self, handler: AuthenticatedHandler, form: dict[str, str | bytes | tuple[str, bytes]]) -> None:
        token = getattr(handler, "_csrf_token", None)
        form_token = str(form.get("csrf_token", ""))
        if not token or form_token != token:
            raise PermissionError("CSRF validation failed.")


def tempfile_dir() -> str:
    return str(Path(os.getenv("TMPDIR", "/tmp")) / "membership-webapp")


def extract_member_id(path: str) -> int | None:
    match = re.search(r"/members/(\d+)/(?:update|usage)$", path)
    return int(match.group(1)) if match else None


def parse_multipart_form(handler: BaseHTTPRequestHandler) -> dict[str, str | bytes | tuple[str, bytes]]:
    content_type = handler.headers.get("Content-Type", "")
    content_length = int(handler.headers.get("Content-Length", "0") or 0)
    if content_type.startswith("multipart/form-data"):
        env = {
            "REQUEST_METHOD": "POST",
            "CONTENT_TYPE": content_type,
            "CONTENT_LENGTH": str(content_length),
        }
        form = cgi.FieldStorage(fp=handler.rfile, headers=handler.headers, environ=env, keep_blank_values=True)
        values: dict[str, str | bytes | tuple[str, bytes]] = {}
        for key in form.keys() if form.list else []:
            field = form[key]
            if isinstance(field, list):
                field = field[0]
            if getattr(field, "filename", None):
                file_data = field.file.read()
                values[key] = (field.filename, file_data)
            else:
                values[key] = field.value
        return values
    raw = handler.rfile.read(content_length).decode("utf-8") if content_length else ""
    parsed = parse_qs(raw)
    return {k: v[-1] for k, v in parsed.items()}


class AuthenticatedHandler(BaseHTTPRequestHandler):
    server_version = f"MembershipWebApp/{APP_VERSION}"

    def _authorized(self) -> bool:
        username = os.getenv("MEMBER_BASIC_AUTH_USER", "")
        password = os.getenv("MEMBER_BASIC_AUTH_PASS", "")
        if not username and not password:
            return True
        header = self.headers.get("Authorization", "")
        if not header.startswith("Basic "):
            return False
        try:
            raw = base64.b64decode(header.split(" ", 1)[1]).decode("utf-8")
        except Exception:
            return False
        candidate_user, _, candidate_pass = raw.partition(":")
        return candidate_user == username and candidate_pass == password

    def _require_auth(self) -> bool:
        if self._authorized():
            return True
        self.send_response(HTTPStatus.UNAUTHORIZED)
        self.send_header("WWW-Authenticate", 'Basic realm="Membership WebApp"')
        self.end_headers()
        self.wfile.write(b"Authentication required")
        return False

    def _secure_headers(self) -> None:
        self.send_header("X-Content-Type-Options", "nosniff")
        self.send_header("X-Frame-Options", "DENY")
        self.send_header("Referrer-Policy", "no-referrer")
        self.send_header("Permissions-Policy", "geolocation=(), microphone=(), camera=()")
        self.send_header("Content-Security-Policy", "default-src 'self'; base-uri 'self'; form-action 'self'; frame-ancestors 'none'")

    def _csrf_cookie_name(self) -> str:
        return "membership_csrf"

    def get_csrf_token(self) -> str:
        cookie = SimpleCookie(self.headers.get("Cookie", ""))
        token = cookie.get(self._csrf_cookie_name())
        if token and token.value:
            self._csrf_token = token.value  # type: ignore[attr-defined]
            return token.value
        value = secrets.token_urlsafe(32)
        self._csrf_token = value  # type: ignore[attr-defined]
        self._set_csrf_cookie = value  # type: ignore[attr-defined]
        return value

    def _send_csrf_cookie_if_needed(self) -> None:
        value = getattr(self, "_set_csrf_cookie", None)
        if not value:
            return
        self.send_header("Set-Cookie", f"{self._csrf_cookie_name()}={value}; HttpOnly; Secure; SameSite=Strict; Path=/")
        delattr(self, "_set_csrf_cookie")

    def do_GET(self) -> None:
        if not self._require_auth():
            return
        self.server.app.handle(self)  # type: ignore[attr-defined]

    def do_POST(self) -> None:
        if not self._require_auth():
            return
        self.server.app.handle(self)  # type: ignore[attr-defined]

    def log_message(self, fmt: str, *args) -> None:
        sys.stdout.write("%s - - [%s] %s\n" % (self.address_string(), self.log_date_time_string(), fmt % args))

    def respond_bytes(self, body: bytes, status: HTTPStatus = HTTPStatus.OK, content_type: str = "text/html; charset=utf-8") -> None:
        self.send_response(status)
        self._secure_headers()
        self._send_csrf_cookie_if_needed()
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def respond_text(self, text: str, status: HTTPStatus = HTTPStatus.OK) -> None:
        self.respond_bytes(text.encode("utf-8"), status=status, content_type="text/plain; charset=utf-8")


class MembershipServer(ThreadingHTTPServer):
    def __init__(self, server_address, RequestHandlerClass, app: MembershipWebApp):
        super().__init__(server_address, RequestHandlerClass)
        self.app = app


def default_db_path() -> Path:
    override = os.getenv("MEMBER_DB_PATH", "").strip()
    if override:
        return Path(override)
    data_dir = app_data_dir_for_platform(sys.platform, Path.home())
    return data_dir / "members.sqlite3"


def default_host() -> str:
    return os.getenv("MEMBER_HOST", "0.0.0.0")


def default_port() -> int:
    return int(os.getenv("MEMBER_PORT", "8000"))


def bootstrap_from_excel_if_present(db: MembershipDatabase, audit: AuditTrail) -> None:
    excel_env = os.getenv("MEMBER_IMPORT_XLSX", "").strip()
    candidates = [Path(excel_env)] if excel_env else []
    candidates.extend(sorted(Path.cwd().glob("*.xlsx")))
    for candidate in candidates:
        if not candidate or not candidate.exists():
            continue
        try:
            result = db.sync_from_workbook(candidate)
            if result.scanned:
                audit.log("bootstrap_sync", {"path": str(candidate), "inserted": result.inserted, "merged": result.merged, "scanned": result.scanned})
            break
        except Exception:
            continue


def main() -> None:
    db = MembershipDatabase(default_db_path())
    audit = AuditTrail(app_data_dir_for_platform(sys.platform, Path.home()))
    if not audit.verify_chain():
        audit.log("audit_chain_failed", {"reason": "startup verification failed"})
    bootstrap_from_excel_if_present(db, audit)
    app = MembershipWebApp(db, audit)
    server = MembershipServer((default_host(), default_port()), AuthenticatedHandler, app)
    print(f"Serving Membership WebApp on http://{default_host()}:{default_port()}")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass
    finally:
        server.server_close()


if __name__ == "__main__":
    main()
